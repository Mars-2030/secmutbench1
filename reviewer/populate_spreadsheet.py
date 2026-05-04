#!/usr/bin/env python3
"""Populate Kill Audit and Test Quality sheets in annotation_spreadsheet.xlsx.

Reads LLM evaluation results and stratified-samples kills (120) and test suites (60)
from the 90 review sample IDs, then writes them into the spreadsheet with COMPLETE
data (code, error messages, prompts, generated tests) — not just reference IDs.

Also generates kill_sample.json and test_sample.json for reviewer reference.
"""

import json
import random
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = Path(__file__).parent.parent
REVIEWER = Path(__file__).parent

# --- Configuration ---

KILL_AUDIT_MODELS = {
    "qwen3-coder:30b": "results/qwen3-coder_30b/baseline_results_20260313_081511.json",
    "gpt-oss-120b": "results/gpt-oss-120b/baseline_results_20260314_183442.json",
    "gpt-5.2": "results/gpt-5.2-2025-12-11/baseline_results_20260314_210150.json",
    "deepseek-v2": "results/deepseek-coder-v2_latest/baseline_results_20260313_192250.json",
}

TEST_QUALITY_MODELS = {
    "qwen3-coder:30b": "results/qwen3-coder_30b/baseline_results_20260313_081511.json",
    "gpt-oss-120b": "results/gpt-oss-120b/baseline_results_20260314_183442.json",
    "deepseek-v2": "results/deepseek-coder-v2_latest/baseline_results_20260313_192250.json",
    "qwen2.5-coder:14b": "results/qwen2.5-coder_14b-instruct/baseline_results_20260314_044805.json",
}

KILL_TYPE_MAP = {
    "semantic": "semantic",
    "assertion_incidental": "incidental",
    "crash": "crash",
    "functional": "functional",
    "other": "other",
}

TARGET_KILLS = 120
TARGET_TESTS = 60
SEED = 42


def load_operator_reference():
    """Parse operator_reference.md into a dict keyed by operator name.

    Returns dict like:
        {"PSQLI": {"description": "...", "target_cwes": "CWE-89",
                    "total_mutants": "80 (80 CWE-specific, 0 generic)",
                    "variants": ["variant1", ...]}, ...}
    """
    import re
    text = (REVIEWER / "operator_reference.md").read_text()
    blocks = re.split(r"^## ", text, flags=re.MULTILINE)[1:]

    operators = {}
    for block in blocks:
        lines = block.strip().split("\n")
        name = lines[0].strip()
        if name == "Summary Table":
            continue
        desc = cwes = total = ""
        variants = []
        in_variants = False
        for line in lines[1:]:
            if line.startswith("**Description**:"):
                desc = line.split(":", 1)[1].strip()
            elif line.startswith("**Target CWEs**:"):
                cwes = line.split(":", 1)[1].strip()
            elif line.startswith("**Total mutants**:"):
                total = line.split(":", 1)[1].strip()
            elif line.startswith("**Mutation variants"):
                in_variants = True
            elif line.startswith("**Example**") or line.startswith("---"):
                in_variants = False
            elif in_variants and line.startswith("- "):
                variants.append(line[2:].strip())
        operators[name] = {
            "description": desc,
            "target_cwes": cwes,
            "total_mutants": total,
            "variants": variants,
        }
    return operators


def format_operator_spec(operator_name, op_ref):
    """Format a single operator's reference info as a readable text block."""
    info = op_ref.get(operator_name)
    if not info:
        return f"{operator_name}: (no reference found)"
    lines = [
        f"{operator_name}",
        f"  {info['description']}",
        f"  Target CWEs: {info['target_cwes']}",
        f"  Total mutants: {info['total_mutants']}",
        f"  Variants: {', '.join(info['variants'])}",
    ]
    return "\n".join(lines)


def format_operators_spec(operator_names, op_ref):
    """Format multiple operators' reference info, one block per operator."""
    ops = [o.strip() for o in operator_names.replace(",", " ").split() if o.strip()]
    unique = list(dict.fromkeys(ops))  # preserve order, dedupe
    return "\n\n".join(format_operator_spec(o, op_ref) for o in unique)


def load_review_ids():
    """Load the 90 review sample IDs and their metadata."""
    with open(REVIEWER / "review_sample.json") as f:
        data = json.load(f)
    return {s["id"]: s for s in data["samples"]}


def load_full_variant(path):
    """Load a result file and return the [full] variant's detailed_results."""
    with open(BASE / path) as f:
        data = json.load(f)
    for r in data["results"]:
        if "[full]" in r["model_name"]:
            return r["detailed_results"]
    raise ValueError(f"No [full] variant in {path}")


# ---------------------------------------------------------------------------
# Kill Audit collection
# ---------------------------------------------------------------------------

def collect_kills(review_ids):
    """Collect all kills from review samples across kill-audit models.

    Each kill record contains the COMPLETE data a reviewer needs:
    secure_code, mutated_code, operator, description, error message, etc.
    """
    kills = []
    for model_name, path in KILL_AUDIT_MODELS.items():
        detailed = load_full_variant(path)
        for dr in detailed:
            sid = dr["sample_id"]
            if sid not in review_ids:
                continue
            for md in dr.get("mutant_details", []):
                if not md.get("killed"):
                    continue
                # Find first failing test error
                error_msg = ""
                failing_test = ""
                for t in md.get("test_results", []):
                    if not t.get("passed") and t.get("error"):
                        error_msg = t["error"]
                        failing_test = t.get("name", "")
                        break

                kill_type = KILL_TYPE_MAP.get(md.get("kill_type", "other"), "other")
                kills.append({
                    "sample_id": sid,
                    "cwe": dr["cwe"],
                    "cwe_name": dr.get("cwe_name", ""),
                    "difficulty": dr.get("difficulty", ""),
                    "entry_point": review_ids[sid].get("entry_point", ""),
                    "operator": md["operator"],
                    "mutant_id": md.get("id", ""),
                    "mutant_category": md.get("mutant_category", ""),
                    "description": md.get("description", ""),
                    "secure_code": dr.get("secure_code", ""),
                    "mutated_code": md.get("mutated_code", ""),
                    "error_message": error_msg,
                    "failing_test": failing_test,
                    "kill_type_raw": md.get("kill_type", ""),
                    "kill_reason": md.get("kill_reason", ""),
                    "heuristic_label": kill_type,
                    "model": model_name,
                })
    return kills


def stratified_sample_kills(kills, target=TARGET_KILLS):
    """Stratified sample: balanced across models and kill types."""
    rng = random.Random(SEED)

    groups = defaultdict(list)
    for k in kills:
        groups[(k["model"], k["heuristic_label"])].append(k)

    for g in groups.values():
        rng.shuffle(g)

    models = list(KILL_AUDIT_MODELS.keys())
    kill_types = ["semantic", "incidental", "crash", "functional", "other"]
    per_model = target // len(models)  # 30

    selected = []
    for model in models:
        model_kills = []
        available_types = [kt for kt in kill_types if groups[(model, kt)]]
        if not available_types:
            continue

        per_type = max(1, per_model // len(available_types))
        remaining = per_model

        for kt in available_types:
            pool = groups[(model, kt)]
            take = min(per_type, len(pool), remaining)
            model_kills.extend(pool[:take])
            remaining -= take

        if remaining > 0:
            for kt in sorted(available_types,
                              key=lambda t: len(groups[(model, t)]),
                              reverse=True):
                pool = groups[(model, kt)]
                already = sum(1 for k in model_kills if k["heuristic_label"] == kt)
                extra = pool[already:]
                take = min(len(extra), remaining)
                model_kills.extend(extra[:take])
                remaining -= take
                if remaining <= 0:
                    break

        selected.extend(model_kills)

    used = {(k["sample_id"], k["mutant_id"], k["model"]) for k in selected}
    remaining = target - len(selected)
    if remaining > 0:
        extras = [k for k in kills
                  if (k["sample_id"], k["mutant_id"], k["model"]) not in used]
        rng.shuffle(extras)
        selected.extend(extras[:remaining])

    return selected[:target]


# ---------------------------------------------------------------------------
# Test Quality collection
# ---------------------------------------------------------------------------

def collect_test_suites(review_ids):
    """Collect valid test suites with COMPLETE data for reviewer."""
    suites = []
    for model_name, path in TEST_QUALITY_MODELS.items():
        detailed = load_full_variant(path)
        for dr in detailed:
            sid = dr["sample_id"]
            if sid not in review_ids:
                continue
            metrics = dr.get("metrics", {})
            if not metrics.get("valid_tests"):
                continue
            if not dr.get("generated_tests"):
                continue
            suites.append({
                "sample_id": sid,
                "cwe": dr["cwe"],
                "cwe_name": dr.get("cwe_name", ""),
                "model": model_name,
                "difficulty": dr.get("difficulty",
                                     review_ids[sid].get("difficulty", "")),
                "entry_point": review_ids[sid].get("entry_point", ""),
                "secure_code": dr.get("secure_code", ""),
                "prompt": dr.get("prompt", ""),
                "generated_tests": dr["generated_tests"],
                "mutation_score": metrics.get("mutation_score", 0),
                "mutants_killed": metrics.get("mutants_killed", 0),
                "mutants_total": metrics.get("mutants_total", 0),
                "tests_count": metrics.get("tests_count", 0),
            })
    return suites


def stratified_sample_tests(suites, review_ids, target=TARGET_TESTS):
    """Stratified sample: 15 per model, balanced 5 easy / 5 medium / 5 hard."""
    rng = random.Random(SEED)

    models = list(TEST_QUALITY_MODELS.keys())
    difficulties = ["easy", "medium", "hard"]
    per_model = target // len(models)
    per_diff = per_model // len(difficulties)

    groups = defaultdict(list)
    for s in suites:
        groups[(s["model"], s["difficulty"])].append(s)
    for g in groups.values():
        rng.shuffle(g)

    selected = []
    for model in models:
        model_suites = []
        remaining = per_model

        for diff in difficulties:
            pool = groups[(model, diff)]
            take = min(per_diff, len(pool), remaining)
            model_suites.extend(pool[:take])
            remaining -= take

        if remaining > 0:
            for diff in difficulties:
                pool = groups[(model, diff)]
                already = sum(1 for s in model_suites if s["difficulty"] == diff)
                extra = pool[already:]
                take = min(len(extra), remaining)
                model_suites.extend(extra[:take])
                remaining -= take
                if remaining <= 0:
                    break

        selected.extend(model_suites)

    return selected[:target]


# ---------------------------------------------------------------------------
# Spreadsheet writing
# ---------------------------------------------------------------------------

def write_spreadsheet(kills, tests, review_samples, op_ref):
    """Write all three sheets into annotation_spreadsheet.xlsx with full data."""
    xlsx_path = REVIEWER / "annotation_spreadsheet.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                              fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")
    code_font = Font(name="Consolas", size=9)

    # ===== Mutation Quality Sheet =====
    ws_mut = wb["Mutation Quality"]

    # Clear all data rows
    for row in range(ws_mut.max_row, 1, -1):
        ws_mut.delete_rows(row)

    # Rewrite header with expanded columns (code + per-mutant details)
    mut_headers = [
        "Sample ID",                                           # A
        "CWE",                                                 # B
        "CWE Name",                                            # C
        "Difficulty",                                          # D
        "Entry Point",                                         # E
        "Source",                                               # F
        "Source Type",                                          # G
        "# Mutants",                                           # H
        "Operators",                                           # I
        "Operator Spec\n(from operator_reference.md)",         # J
        "Secure Code",                                         # K
        "Insecure Code",                                       # L
        "Mutant Details\n(ID | Operator | Category | Description)",  # M
        "Mutant Code\n(all mutants, separated by dividers)",   # N
        "C1: Secure Code\n(YES/PARTIAL/NO)",                   # O
        "C2: Mutant Realism\n(REALISTIC/CONTRIVED/INVALID/EQUIVALENT)",  # P
        "C3: CWE Label\n(CORRECT/RELATED/INCORRECT)",          # Q
        "C4: Operator Alignment\n(ALIGNED/PARTIAL/MISALIGNED)",  # R
        "C5: Difficulty Rating\n(APPROPRIATE/TOO EASY/TOO HARD)",  # S
        "Comments",                                            # T
    ]
    for col_idx, hdr in enumerate(mut_headers, start=1):
        cell = ws_mut.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Write mutation quality data rows
    code_cols_mut = {10, 11, 12, 13, 14}  # J-N = operator spec/code/mutant columns
    for i, sample in enumerate(review_samples, start=1):
        row = i + 1
        mutants = sample.get("mutants", [])
        operators = ", ".join(sorted(set(
            m["operator"] for m in mutants
        )))
        operator_spec = format_operators_spec(operators, op_ref)

        # Build per-mutant details text
        mutant_details_lines = []
        mutant_code_parts = []
        for j, m in enumerate(mutants, start=1):
            mid = m.get("id", "?")
            op = m.get("operator", "?")
            cat = m.get("mutant_category", "?")
            desc = m.get("description", "")
            vtype = m.get("variant_type", "")
            mutant_details_lines.append(
                f"[{j}] {mid} | {op} | {cat} | {desc}"
                + (f" ({vtype})" if vtype else "")
            )
            mutant_code_parts.append(
                f"--- Mutant {j}: {mid} ({op}) ---\n"
                + m.get("mutated_code", "")
            )

        mutant_details_text = "\n".join(mutant_details_lines)
        mutant_code_text = "\n\n".join(mutant_code_parts)

        values = [
            sample["id"],                          # A
            sample["cwe"],                         # B
            sample.get("cwe_name", ""),             # C
            sample.get("difficulty", ""),           # D
            sample.get("entry_point", ""),          # E
            sample.get("source", ""),               # F
            sample.get("source_type", ""),          # G
            len(mutants),                           # H
            operators,                              # I
            operator_spec,                          # J
            sample.get("secure_code", ""),          # K
            sample.get("insecure_code", ""),        # L
            mutant_details_text,                    # M
            mutant_code_text,                       # N
            None,                                   # O - reviewer
            None,                                   # P - reviewer
            None,                                   # Q - reviewer
            None,                                   # R - reviewer
            None,                                   # S - reviewer
            None,                                   # T - reviewer
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_mut.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in code_cols_mut:
                cell.alignment = wrap_align
                cell.font = code_font
            else:
                cell.alignment = top_align

    # Column widths for Mutation Quality
    mut_col_widths = {
        "A": 14, "B": 10, "C": 22, "D": 10, "E": 22, "F": 14,
        "G": 12, "H": 10, "I": 22, "J": 50, "K": 55, "L": 55,
        "M": 55, "N": 60, "O": 16, "P": 18, "Q": 16, "R": 18,
        "S": 18, "T": 20,
    }
    for col_letter, width in mut_col_widths.items():
        ws_mut.column_dimensions[col_letter].width = width

    # ===== Kill Audit Sheet =====
    # Replace the entire sheet to add new columns for complete data
    ws_kill = wb["Kill Audit"]

    # Clear all data rows
    for row in range(ws_kill.max_row, 1, -1):
        ws_kill.delete_rows(row)

    # Rewrite header with expanded columns
    kill_headers = [
        "Kill ID",                                          # A
        "Sample ID",                                        # B
        "CWE",                                              # C
        "CWE Name",                                         # D
        "Difficulty",                                       # E
        "Entry Point",                                      # F
        "Operator",                                         # G
        "Operator Spec\n(from operator_reference.md)",      # H
        "Mutant Category",                                  # I
        "Mutant Description",                               # J
        "Secure Code",                                      # K
        "Mutated Code",                                     # L
        "Error Message",                                    # M
        "Failing Test Name",                                # N
        "Model",                                            # O
        "Heuristic Label\n(semantic/incidental/crash/other)",  # P
        "Your Classification\n(semantic/incidental/crash/other)",  # Q
        "Agree?\n(AGREE/DISAGREE)",                         # R
        "If Disagree: Reasoning",                           # S
        "Confidence\n(HIGH/MEDIUM/LOW)",                    # T
    ]
    for col_idx, hdr in enumerate(kill_headers, start=1):
        cell = ws_kill.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Write kill data rows
    code_cols_kill = {8, 11, 12, 13}  # H=op_spec, K=secure, L=mutated, M=error
    for i, kill in enumerate(kills, start=1):
        row = i + 1
        kill_id = f"K{i:03d}"
        op_spec = format_operator_spec(kill["operator"], op_ref)
        values = [
            kill_id,                    # A
            kill["sample_id"],          # B
            kill["cwe"],                # C
            kill["cwe_name"],           # D
            kill["difficulty"],         # E
            kill["entry_point"],        # F
            kill["operator"],           # G
            op_spec,                    # H
            kill["mutant_category"],    # I
            kill["description"],        # J
            kill["secure_code"],        # K
            kill["mutated_code"],       # L
            kill["error_message"],      # M
            kill["failing_test"],       # N
            kill["model"],              # O
            kill["heuristic_label"],    # P
            None,                       # Q - reviewer
            None,                       # R - reviewer
            None,                       # S - reviewer
            None,                       # T - reviewer
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_kill.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in code_cols_kill:
                cell.alignment = wrap_align
                cell.font = code_font
            elif col_idx == 10:  # description
                cell.alignment = wrap_align
            else:
                cell.alignment = top_align

    # Column widths for Kill Audit
    kill_col_widths = {
        "A": 8, "B": 14, "C": 10, "D": 22, "E": 10, "F": 20,
        "G": 14, "H": 50, "I": 14, "J": 40, "K": 55, "L": 55,
        "M": 55, "N": 30, "O": 18, "P": 16, "Q": 18, "R": 10,
        "S": 30, "T": 12,
    }
    for col_letter, width in kill_col_widths.items():
        ws_kill.column_dimensions[col_letter].width = width

    # ===== Test Quality Sheet =====
    ws_test = wb["Test Quality"]

    for row in range(ws_test.max_row, 1, -1):
        ws_test.delete_rows(row)

    # Rewrite header with expanded columns
    test_headers = [
        "Sample ID",                                           # A
        "CWE",                                                 # B
        "CWE Name",                                            # C
        "Model",                                               # D
        "Difficulty",                                          # E
        "Entry Point",                                         # F
        "Secure Code",                                         # G
        "Prompt Given to Model",                               # H
        "Generated Tests",                                     # I
        "Mutation Score",                                       # J
        "Mutants Killed / Total",                              # K
        "# Tests",                                             # L
        "C1: Security Awareness\n(HIGH/MEDIUM/LOW/NONE)",      # M
        "C2: Attack Vector Coverage\n(COMPREHENSIVE/PARTIAL/MISSING)",  # N
        "C3: Assertion Quality\n(STRONG/ADEQUATE/WEAK/BROKEN)",  # O
        "C4: Test Correctness\n(CORRECT/PARTIAL/INCORRECT/UNCERTAIN)",  # P
        "C5: Mock/Import Awareness\n(COMPATIBLE/MOCK-DEPENDENT/INCOMPATIBLE)",  # Q
        "Free-Text Assessment\n(1-3 sentences)",               # R
        "Comments",                                            # S
    ]
    for col_idx, hdr in enumerate(test_headers, start=1):
        cell = ws_test.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Write test data rows
    code_cols_test = {7, 8, 9}  # G=secure_code, H=prompt, I=generated_tests
    for i, test in enumerate(tests, start=1):
        row = i + 1
        ms = test["mutation_score"]
        kill_frac = f"{test['mutants_killed']}/{test['mutants_total']}"
        values = [
            test["sample_id"],          # A
            test["cwe"],                # B
            test["cwe_name"],           # C
            test["model"],              # D
            test["difficulty"],         # E
            test["entry_point"],        # F
            test["secure_code"],        # G
            test["prompt"],             # H
            test["generated_tests"],    # I
            round(ms, 3) if ms is not None else 0,  # J
            kill_frac,                  # K
            test["tests_count"],        # L
            None,                       # M - reviewer
            None,                       # N - reviewer
            None,                       # O - reviewer
            None,                       # P - reviewer
            None,                       # Q - reviewer
            None,                       # R - reviewer
            None,                       # S - reviewer
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_test.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in code_cols_test:
                cell.alignment = wrap_align
                cell.font = code_font
            else:
                cell.alignment = top_align

    # Column widths for Test Quality
    test_col_widths = {
        "A": 14, "B": 10, "C": 22, "D": 22, "E": 10, "F": 20,
        "G": 55, "H": 60, "I": 60, "J": 12, "K": 14, "L": 8,
        "M": 18, "N": 22, "O": 18, "P": 18, "Q": 22, "R": 35, "S": 20,
    }
    for col_letter, width in test_col_widths.items():
        ws_test.column_dimensions[col_letter].width = width

    wb.save(xlsx_path)
    print(f"Saved spreadsheet: {xlsx_path}")


# ---------------------------------------------------------------------------
# JSON sample files (for reviewer reference alongside the spreadsheet)
# ---------------------------------------------------------------------------

def write_json_samples(kills, tests):
    """Write kill_sample.json and test_sample.json for reviewer."""
    # kill_sample.json
    kill_records = []
    for i, k in enumerate(kills, start=1):
        kill_records.append({
            "kill_id": f"K{i:03d}",
            "sample_id": k["sample_id"],
            "cwe": k["cwe"],
            "cwe_name": k["cwe_name"],
            "difficulty": k["difficulty"],
            "entry_point": k["entry_point"],
            "operator": k["operator"],
            "mutant_id": k["mutant_id"],
            "mutant_category": k["mutant_category"],
            "description": k["description"],
            "secure_code": k["secure_code"],
            "mutated_code": k["mutated_code"],
            "error_message": k["error_message"],
            "failing_test": k["failing_test"],
            "model": k["model"],
            "heuristic_label": k["heuristic_label"],
        })

    kill_path = REVIEWER / "kill_sample.json"
    with open(kill_path, "w") as f:
        json.dump({
            "description": "Stratified kill sample: 120 kills for expert audit",
            "total_kills": len(kill_records),
            "kills": kill_records,
        }, f, indent=2)
    print(f"Saved: {kill_path}")

    # test_sample.json
    test_records = []
    for t in tests:
        test_records.append({
            "sample_id": t["sample_id"],
            "cwe": t["cwe"],
            "cwe_name": t["cwe_name"],
            "model": t["model"],
            "difficulty": t["difficulty"],
            "entry_point": t["entry_point"],
            "secure_code": t["secure_code"],
            "prompt": t["prompt"],
            "generated_tests": t["generated_tests"],
            "mutation_score": round(t["mutation_score"], 4) if t["mutation_score"] is not None else 0,
            "mutants_killed": t["mutants_killed"],
            "mutants_total": t["mutants_total"],
            "tests_count": t["tests_count"],
        })

    test_path = REVIEWER / "test_sample.json"
    with open(test_path, "w") as f:
        json.dump({
            "description": "Stratified test quality sample: 60 suites for expert audit",
            "total_suites": len(test_records),
            "suites": test_records,
        }, f, indent=2)
    print(f"Saved: {test_path}")


# ---------------------------------------------------------------------------
# Stats & main
# ---------------------------------------------------------------------------

def print_stats(kills, tests):
    """Print summary statistics for verification."""
    print("\n=== Kill Audit Stats ===")
    print(f"Total kills: {len(kills)}")
    print(f"By model: {dict(Counter(k['model'] for k in kills))}")
    print(f"By kill type: {dict(Counter(k['heuristic_label'] for k in kills))}")
    cwe_dist = Counter(k["cwe"] for k in kills)
    print(f"CWE coverage: {len(cwe_dist)} CWEs")
    print(f"By CWE: {dict(sorted(cwe_dist.items()))}")
    print(f"Unique sample IDs: {len(set(k['sample_id'] for k in kills))}")
    # Verify all have code
    no_secure = sum(1 for k in kills if not k["secure_code"])
    no_mutated = sum(1 for k in kills if not k["mutated_code"])
    no_error = sum(1 for k in kills if not k["error_message"])
    print(f"Data completeness: {no_secure} missing secure_code, "
          f"{no_mutated} missing mutated_code, {no_error} missing error_message")

    print("\n=== Test Quality Stats ===")
    print(f"Total test suites: {len(tests)}")
    print(f"By model: {dict(Counter(t['model'] for t in tests))}")
    print(f"By difficulty: {dict(Counter(t['difficulty'] for t in tests))}")
    cwe_dist = Counter(t["cwe"] for t in tests)
    print(f"CWE coverage: {len(cwe_dist)} CWEs")
    for model in TEST_QUALITY_MODELS:
        mdiff = Counter(t["difficulty"] for t in tests if t["model"] == model)
        print(f"  {model}: {dict(mdiff)}")
    no_code = sum(1 for t in tests if not t["secure_code"])
    no_prompt = sum(1 for t in tests if not t["prompt"])
    no_gen = sum(1 for t in tests if not t["generated_tests"])
    print(f"Data completeness: {no_code} missing secure_code, "
          f"{no_prompt} missing prompt, {no_gen} missing generated_tests")


def main():
    review_ids = load_review_ids()
    print(f"Loaded {len(review_ids)} review sample IDs")

    op_ref = load_operator_reference()
    print(f"Loaded {len(op_ref)} operator specs from operator_reference.md")

    # Load ordered review samples list for Mutation Quality sheet
    with open(REVIEWER / "review_sample.json") as f:
        review_samples = json.load(f)["samples"]

    # Kill Audit
    all_kills = collect_kills(review_ids)
    print(f"Total available kills from review samples: {len(all_kills)}")
    kills = stratified_sample_kills(all_kills, TARGET_KILLS)

    # Test Quality
    all_suites = collect_test_suites(review_ids)
    print(f"Total available test suites from review samples: {len(all_suites)}")
    tests = stratified_sample_tests(all_suites, review_ids, TARGET_TESTS)

    print_stats(kills, tests)

    write_spreadsheet(kills, tests, review_samples, op_ref)
    write_json_samples(kills, tests)
    print("\nDone! Spreadsheet and JSON files updated.")


if __name__ == "__main__":
    main()
